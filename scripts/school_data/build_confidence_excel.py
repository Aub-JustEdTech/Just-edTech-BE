"""Build a client-facing Excel workbook from schema-driven crawl JSON results.

Groups discovered "data pages" (board minutes / agendas / policy documents pages)
by confidence score into separate sheets, so a non-technical client can review
and confirm the correct URL for each school.

Usage:
    poetry run python scripts/school_data/build_confidence_excel.py \
        --input scripts/school_data/output/schema_crawl_2026-08-05/schema_crawl_results_batch_001_100.json \
                scripts/school_data/output/schema_crawl_2026-08-05/schema_crawl_results_batch_101_200.json \
                scripts/school_data/output/schema_crawl_2026-08-05/schema_crawl_results_batch_201_300.json \
                scripts/school_data/output/schema_crawl_2026-08-05/schema_crawl_results_batch_301_394.json \
        --output scripts/school_data/output/school_data_review.xlsx

    # Sample of the first 10 schools only:
    poetry run python scripts/school_data/build_confidence_excel.py \
        --input scripts/school_data/output/schema_crawl_2026-08-05/schema_crawl_results_batch_001_100.json \
        --output scripts/school_data/output/school_data_review_SAMPLE.xlsx \
        --sample 10
"""

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MAX_ALT_OPTIONS = 20  # width reserved for hidden "possible URL" helper columns
VALID_YEARS = {2023, 2024, 2025, 2026}


@dataclass
class DataPageRow:
    school_name: str
    website: str
    org_code: str
    url: str
    confidence: float
    data_type: str
    alt_urls: list[str] = field(default_factory=list)
    years_available: list[int] = field(default_factory=list)


@dataclass
class NoDataSchool:
    school_name: str
    website: str
    org_code: str
    pages_crawled: int


@dataclass
class ErrorSchool:
    school_name: str
    website: str
    org_code: str
    has_data: bool
    errors: list[str]


def load_schools(input_files: list[Path]) -> list[dict]:
    schools: list[dict] = []
    seen_org_codes: set[str] = set()
    for f in input_files:
        data = json.loads(f.read_text())
        for school in data:
            org_code = school.get("org_code")
            if org_code in seen_org_codes:
                continue
            seen_org_codes.add(org_code)
            schools.append(school)
    return schools


def bucket_for(confidence: float) -> str:
    if confidence >= 1.0:
        return "Confidence 1.0"
    if confidence >= 0.7:
        return "Confidence 0.7-0.9"
    if confidence >= 0.5:
        return "Confidence 0.5-0.7"
    return "Below 0.5 Confidence"


BUCKET_ORDER = [
    "Confidence 1.0",
    "Confidence 0.7-0.9",
    "Confidence 0.5-0.7",
    "Below 0.5 Confidence",
]


def build_rows(
    schools: list[dict],
) -> tuple[dict[str, list[DataPageRow]], list[DataPageRow], list[NoDataSchool], list[ErrorSchool]]:
    buckets: dict[str, list[DataPageRow]] = {b: [] for b in BUCKET_ORDER}
    outdated: list[DataPageRow] = []
    no_data: list[NoDataSchool] = []
    errored: list[ErrorSchool] = []

    for school in schools:
        name = school.get("name", "")
        website = school.get("website", "")
        org_code = school.get("org_code", "")
        data_pages = school.get("data_pages") or []
        errors = school.get("errors") or []

        if errors:
            errored.append(
                ErrorSchool(
                    school_name=name,
                    website=website,
                    org_code=org_code,
                    has_data=bool(data_pages),
                    errors=[str(e) for e in errors],
                )
            )

        if not data_pages:
            if not errors:
                no_data.append(
                    NoDataSchool(
                        school_name=name,
                        website=website,
                        org_code=org_code,
                        pages_crawled=school.get("pages_crawled", 0),
                    )
                )
            continue

        for dp in data_pages:
            info = dp.get("data_page_info") or {}
            confidence = info.get("confidence")
            if confidence is None:
                continue
            main_url = dp.get("url", "")
            alt_urls = [main_url]
            for rel in dp.get("possible_relevant_pages") or []:
                rel_url = rel.get("url")
                if rel_url and rel_url not in alt_urls:
                    alt_urls.append(rel_url)

            years_available = info.get("data_years_available") or []

            row = DataPageRow(
                school_name=name,
                website=website,
                org_code=org_code,
                url=main_url,
                confidence=confidence,
                data_type=info.get("data_type", "unknown"),
                alt_urls=alt_urls[:MAX_ALT_OPTIONS],
                years_available=years_available,
            )

            # Pages entirely outside the client's valid year range (2023-2026) are
            # stale archives - route them to a separate "outdated" sheet instead of
            # a confidence bucket, even if the model was fully confident it's a data page.
            if years_available and not any(y in VALID_YEARS for y in years_available):
                outdated.append(row)
            else:
                buckets[bucket_for(confidence)].append(row)

    for rows in buckets.values():
        rows.sort(key=lambda r: (-r.confidence, r.school_name.lower()))
    outdated.sort(key=lambda r: r.school_name.lower())
    no_data.sort(key=lambda r: r.school_name.lower())
    errored.sort(key=lambda r: r.school_name.lower())

    return buckets, outdated, no_data, errored


def style_header(ws: Worksheet, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def write_bucket_sheet(
    wb: Workbook, bucket_name: str, rows: list[DataPageRow], show_years_note: bool = False
) -> None:
    ws = wb.create_sheet(title=bucket_name[:31])
    headers = [
        "School Name",
        "School Website",
        "Data Page URL (found)",
        "Confidence",
        "Data Type",
        "Confirmed URL",
        "Client Confirmed?",
    ]
    if show_years_note:
        headers.append("Years Available")
    style_header(ws, headers)

    confirmed_dv = DataValidation(
        type="list", formula1='"Yes,No,Needs Correction"', allow_blank=True, showDropDown=False
    )
    ws.add_data_validation(confirmed_dv)

    alt_start_col = len(headers) + 2  # leave one blank spacer column, then hidden helper columns
    alt_end_col = alt_start_col + MAX_ALT_OPTIONS - 1

    for i, row in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=row.school_name)
        ws.cell(row=i, column=2, value=row.website)
        ws.cell(row=i, column=3, value=row.url)
        conf_cell = ws.cell(row=i, column=4, value=row.confidence)
        conf_cell.number_format = "0.00"
        ws.cell(row=i, column=5, value=row.data_type)
        confirmed_url_cell = ws.cell(row=i, column=6, value=row.url)
        ws.cell(row=i, column=7)
        if show_years_note:
            ws.cell(row=i, column=8, value=", ".join(str(y) for y in row.years_available))

        for j, alt_url in enumerate(row.alt_urls):
            ws.cell(row=i, column=alt_start_col + j, value=alt_url)

        url_range = (
            f"${get_column_letter(alt_start_col)}{i}:${get_column_letter(alt_end_col)}{i}"
        )
        url_dv = DataValidation(type="list", formula1=url_range, allow_blank=True, showDropDown=False)
        ws.add_data_validation(url_dv)
        url_dv.add(confirmed_url_cell)
        confirmed_dv.add(ws.cell(row=i, column=7))

    # Hide the helper columns used for the per-row URL dropdown options
    for col in range(alt_start_col, alt_end_col + 1):
        ws.column_dimensions[get_column_letter(col)].hidden = True

    widths = [30, 32, 45, 11, 16, 45, 18]
    if show_years_note:
        widths.append(16)
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def write_no_data_sheet(wb: Workbook, rows: list[NoDataSchool]) -> None:
    ws = wb.create_sheet(title="No Data Found")
    headers = ["School Name", "School Website", "Org Code", "Pages Crawled", "Client Notes"]
    style_header(ws, headers)
    for i, row in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=row.school_name)
        ws.cell(row=i, column=2, value=row.website)
        ws.cell(row=i, column=3, value=row.org_code)
        ws.cell(row=i, column=4, value=row.pages_crawled)
    widths = [30, 32, 14, 14, 40]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def write_error_sheet(wb: Workbook, rows: list[ErrorSchool]) -> None:
    ws = wb.create_sheet(title="Crawl Errors")
    headers = ["School Name", "School Website", "Org Code", "Data Found Anyway?", "Error Details"]
    style_header(ws, headers)
    for i, row in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=row.school_name)
        ws.cell(row=i, column=2, value=row.website)
        ws.cell(row=i, column=3, value=row.org_code)
        ws.cell(row=i, column=4, value="Yes" if row.has_data else "No")
        ws.cell(row=i, column=5, value="; ".join(row.errors))
    widths = [30, 32, 14, 16, 60]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def write_summary_sheet(
    wb: Workbook,
    total_schools: int,
    buckets: dict[str, list[DataPageRow]],
    outdated: list[DataPageRow],
    no_data: list[NoDataSchool],
    errored: list[ErrorSchool],
) -> None:
    ws = wb.create_sheet(title="Summary", index=0)
    ws["A1"] = "School Website Data Review - Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Total schools processed", total_schools])
    ws.append([])
    ws.append(["Sheet", "Row Count", "What it means"])
    for cell in ws[4]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    unique_schools_per_bucket = {
        name: len({r.school_name for r in rows}) for name, rows in buckets.items()
    }
    descriptions = {
        "Confidence 1.0": "Data page found with maximum confidence - very likely correct.",
        "Confidence 0.7-0.9": "Data page found with high confidence - please double check.",
        "Confidence 0.5-0.7": "Data page found with medium confidence - please verify carefully.",
        "Below 0.5 Confidence": "Data page found with low confidence - likely needs correction.",
    }
    for bucket_name in BUCKET_ORDER:
        ws.append(
            [
                bucket_name,
                f"{len(buckets[bucket_name])} pages / {unique_schools_per_bucket[bucket_name]} schools",
                descriptions[bucket_name],
            ]
        )
    ws.append(
        [
            "Outdated / Out of Range",
            f"{len(outdated)} pages",
            "Data page found, but all available years are before 2023 - too old to be useful.",
        ]
    )
    ws.append(["No Data Found", len(no_data), "No relevant data page could be located for these schools."])
    ws.append(["Crawl Errors", len(errored), "The crawler hit an error for these schools (may still have partial data)."])

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 70


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", nargs="+", required=True, type=Path, help="Input JSON batch file(s)")
    parser.add_argument("--output", required=True, type=Path, help="Output .xlsx path")
    parser.add_argument("--sample", type=int, default=None, help="Only take the first N schools (for a sample run)")
    args = parser.parse_args()

    schools = load_schools(args.input)
    if args.sample:
        schools = schools[: args.sample]

    buckets, outdated, no_data, errored = build_rows(schools)

    wb = Workbook()
    wb.remove(wb.active)

    write_summary_sheet(wb, len(schools), buckets, outdated, no_data, errored)
    for bucket_name in BUCKET_ORDER:
        write_bucket_sheet(wb, bucket_name, buckets[bucket_name])
    write_bucket_sheet(wb, "Outdated - Out of Range", outdated, show_years_note=True)
    write_no_data_sheet(wb, no_data)
    write_error_sheet(wb, errored)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Wrote {args.output} ({len(schools)} schools)")


if __name__ == "__main__":
    main()
