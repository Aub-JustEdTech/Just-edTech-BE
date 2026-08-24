#!/usr/bin/env python3
"""
Run the schema-driven LLM crawler over unique schools listed in a spreadsheet
and write summary results to a separate sheet in the same workbook.

Reads ``School Name`` + ``School Website`` columns (tolerates minor header
variants). Deduplicates by normalised (name, website) before crawling.

Crawler settings (fixed for this batch runner):
  - max_depth=4
  - skip_archival=False  (do not skip archive pages the LLM marks is_archive)
  - max_pages_limit_reached surfaced on each row (LLM page-budget flag)

Usage:
    python scripts/school_data/run_schema_crawl_from_xlsx.py \\
        --input "scripts/school_data/Untitled spreadsheet.xlsx"
    python scripts/school_data/run_schema_crawl_from_xlsx.py --limit 5
    python scripts/school_data/run_schema_crawl_from_xlsx.py --resume
    python scripts/school_data/run_schema_crawl_from_xlsx.py \\
        --retry-under-explored --max-pages 30 --resume
"""

from __future__ import annotations

import argparse
import asyncio
import json
import re
import sys
import time
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.config import settings
from app.services.web_scraper.page_classifier import PageClassifier
from app.services.web_scraper.schema_driven_crawler import (
    SchemaDrivenCrawler,
    _keyword_boost,
)
from app.services.web_scraper.schema_driven_school_scraper_service import (
    _HUB_FALLBACK_SCORE,
)

DEFAULT_INPUT = Path(__file__).parent / "Untitled spreadsheet.xlsx"
RESULTS_SHEET = "Schema Crawl Results"
DEFAULT_JSON_OUT = Path(__file__).parent / "output" / "schema_crawl_xlsx_results.json"
DEFAULT_RETRY_JSON_OUT = (
    Path(__file__).parent / "output" / "schema_crawl_xlsx_retry_progress.json"
)
DEFAULT_RETRY_MAX_PAGES = 30
DEFAULT_RETRY_CANDIDATE_THRESHOLD = 3

RESULT_HEADERS = [
    "School Name",
    "School Website",
    "Status",
    "Pages Crawled",
    "LLM Calls",
    "Max Budget Reached",
    "Candidate Count",
    "Retried With Higher Budget",
    "Top Candidate URLs",
    "Candidate Details (JSON)",
    "Error",
    "Elapsed Seconds",
    "Crawled At (UTC)",
]

_NAME_ALIASES = ("school name", "name", "district", "district name")
_WEBSITE_ALIASES = ("school website", "website", "url", "homepage")


def _norm_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _match_col(headers: list[str], aliases: tuple[str, ...]) -> int | None:
    for idx, header in enumerate(headers):
        if header in aliases:
            return idx
    return None


def load_schools_from_xlsx(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError(f"Spreadsheet {path} is empty.")

    headers = [_norm_header(c) for c in rows[0]]
    name_idx = _match_col(headers, _NAME_ALIASES)
    website_idx = _match_col(headers, _WEBSITE_ALIASES)
    if name_idx is None or website_idx is None:
        raise ValueError(
            f"Could not find School Name / School Website columns. Headers: {rows[0]}"
        )

    seen: set[tuple[str, str]] = set()
    schools: list[dict[str, str]] = []
    for row in rows[1:]:
        if not row:
            continue
        raw_name = row[name_idx]
        if raw_name is None or not str(raw_name).strip():
            continue
        name = str(raw_name).strip()
        website = str(row[website_idx] or "").strip().rstrip("/")
        key = (name.lower(), website.lower())
        if key in seen:
            continue
        seen.add(key)
        schools.append({"name": name, "website": website})
    return schools


def _school_key(record: dict[str, str]) -> str:
    return f"{record['name'].lower()}|{record['website'].lower()}"


def load_completed_keys(path: Path) -> set[str]:
    if not path.exists():
        return set()
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return set()
    return {
        _school_key({"name": r.get("name", ""), "website": r.get("website", "")})
        for r in data.get("districts", [])
        if r.get("name")
    }


def load_results(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    try:
        return json.loads(path.read_text(encoding="utf-8")).get("districts", [])
    except json.JSONDecodeError:
        return []


def select_under_explored(
    results: list[dict[str, Any]],
    *,
    candidate_threshold: int,
) -> list[dict[str, str]]:
    """Schools that hit the page budget but returned fewer than threshold candidates."""
    selected: list[dict[str, str]] = []
    for row in results:
        if not row.get("max_pages_limit_reached"):
            continue
        if row.get("candidate_count", 0) >= candidate_threshold:
            continue
        name = (row.get("name") or "").strip()
        website = (row.get("website") or "").strip().rstrip("/")
        if not name:
            continue
        selected.append({"name": name, "website": website})
    return selected


def merge_result_update(
    existing: list[dict[str, Any]],
    update: dict[str, Any],
    *,
    order: dict[str, int],
) -> list[dict[str, Any]]:
    key = _school_key({"name": update.get("name", ""), "website": update.get("website", "")})
    by_key = {
        _school_key({"name": r.get("name", ""), "website": r.get("website", "")}): r
        for r in existing
    }
    by_key[key] = update
    return sorted(
        by_key.values(),
        key=lambda r: order.get(
            _school_key({"name": r.get("name", ""), "website": r.get("website", "")}),
            10_000,
        ),
    )


def _candidate_urls(candidates: list[dict[str, Any]]) -> str:
    return " | ".join(c.get("url", "") for c in candidates if c.get("url"))


def _candidate_json(candidates: list[dict[str, Any]]) -> str:
    slim = [
        {
            "url": c.get("url"),
            "score": c.get("score"),
            "data_type": c.get("data_type"),
            "is_archive": c.get("is_archive"),
            "data_years_available": c.get("data_years_available"),
        }
        for c in candidates
    ]
    return json.dumps(slim, ensure_ascii=False)


def _format_candidates(crawl_result: Any, *, max_candidates: int) -> list[dict[str, Any]]:
    """Mirror SchemaDrivenSchoolScraperService candidate shaping."""
    candidates: list[dict[str, Any]] = []
    seen: set[str] = set()
    for page in crawl_result.data_pages:
        url = (page.url or "").strip()
        if not url or url in seen:
            continue
        seen.add(url)
        info = page.data_page_info
        candidates.append(
            {
                "url": url,
                "matched_keywords": [],
                "score": int((info.confidence * 100) if info else 100),
                "data_type": info.data_type if info else None,
                "is_archive": bool(info.is_archive) if info else False,
                "data_years_available": list(info.data_years_available) if info else [],
            }
        )

    if len(candidates) < 1:
        for page in crawl_result.visited_pages:
            url = (page.url or "").strip()
            if not url or url in seen:
                continue
            if not page.has_data_links:
                continue
            if _keyword_boost(url) <= 0:
                continue
            seen.add(url)
            candidates.append(
                {
                    "url": url,
                    "matched_keywords": [],
                    "score": _HUB_FALLBACK_SCORE,
                    "data_type": "hub",
                    "is_archive": False,
                    "data_years_available": [],
                }
            )

    candidates.sort(key=lambda c: c["score"], reverse=True)
    return candidates[:max_candidates]


async def crawl_one(
    record: dict[str, str],
    classifier: PageClassifier,
    *,
    max_depth: int,
    skip_archival: bool,
    max_pages: int,
    max_candidates: int,
) -> dict[str, Any]:
    name = record["name"]
    website = record["website"]
    started = time.monotonic()
    crawled_at = datetime.now(UTC).isoformat()

    result: dict[str, Any] = {
        "name": name,
        "website": website,
        "status": "pending",
        "pages_crawled": 0,
        "llm_calls": 0,
        "max_pages_limit_reached": False,
        "candidate_count": 0,
        "candidates": [],
        "error": None,
        "elapsed_seconds": 0.0,
        "crawled_at": crawled_at,
    }

    if not website:
        result["status"] = "missing_website"
        result["error"] = "missing website"
        result["elapsed_seconds"] = round(time.monotonic() - started, 1)
        return result

    crawler = SchemaDrivenCrawler(
        classifier=classifier,
        max_pages=max_pages,
        max_depth=max_depth,
        skip_archival=skip_archival,
    )
    try:
        crawl_result = await crawler.crawl(website)
        candidates = _format_candidates(crawl_result, max_candidates=max_candidates)
        result["pages_crawled"] = crawl_result.pages_crawled
        result["llm_calls"] = crawl_result.llm_calls
        result["max_pages_limit_reached"] = bool(crawl_result.max_pages_limit_reached)
        result["candidates"] = candidates
        result["candidate_count"] = len(candidates)
        result["status"] = "ok"
        if crawl_result.errors:
            result["error"] = "; ".join(crawl_result.errors)
    except Exception as exc:  # noqa: BLE001
        result["status"] = "failed"
        result["error"] = f"{type(exc).__name__}: {exc}"
    finally:
        await crawler.close()
        result["elapsed_seconds"] = round(time.monotonic() - started, 1)

    return result


def _result_to_row(result: dict[str, Any]) -> list[Any]:
    candidates = result.get("candidates") or []
    return [
        result.get("name"),
        result.get("website"),
        result.get("status"),
        result.get("pages_crawled", 0),
        result.get("llm_calls", 0),
        bool(result.get("max_pages_limit_reached")),
        result.get("candidate_count", 0),
        bool(result.get("retry_pass")),
        _candidate_urls(candidates),
        _candidate_json(candidates),
        result.get("error"),
        result.get("elapsed_seconds", 0.0),
        result.get("crawled_at"),
    ]


def write_results_sheet(path: Path, results: list[dict[str, Any]]) -> None:
    wb = load_workbook(path)
    if RESULTS_SHEET in wb.sheetnames:
        del wb[RESULTS_SHEET]
    ws = wb.create_sheet(RESULTS_SHEET)
    ws.append(RESULT_HEADERS)
    for result in results:
        ws.append(_result_to_row(result))
    wb.save(path)
    wb.close()


def write_json_report(path: Path, results: list[dict[str, Any]], meta: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "generated_at": datetime.now(UTC).isoformat(),
        **meta,
        "districts": results,
    }
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


async def run(args: argparse.Namespace) -> None:
    input_path: Path = args.input
    if not input_path.exists():
        print(f"Input spreadsheet not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    schools = load_schools_from_xlsx(input_path)
    order = {_school_key(s): i for i, s in enumerate(schools)}
    max_pages = (
        args.max_pages
        if args.max_pages is not None
        else (
            DEFAULT_RETRY_MAX_PAGES
            if args.retry_under_explored
            else settings.SCHOOL_SCRAPER_LLM_MAX_PAGES
        )
    )
    max_candidates = settings.SCHOOL_SCRAPER_MAX_CANDIDATES

    if args.retry_under_explored:
        if not args.json_out.exists():
            print(f"Results JSON not found for retry: {args.json_out}", file=sys.stderr)
            sys.exit(1)
        existing_results = load_results(args.json_out)
        if not existing_results:
            print(f"No districts in results JSON: {args.json_out}", file=sys.stderr)
            sys.exit(1)
        retry_targets = select_under_explored(
            existing_results,
            candidate_threshold=args.candidate_threshold,
        )
        retried = load_completed_keys(args.retry_json_out) if args.resume else set()
        pending = [s for s in retry_targets if _school_key(s) not in retried]
    else:
        existing_results = load_results(args.json_out) if args.resume else []
        completed = load_completed_keys(args.json_out) if args.resume else set()
        pending = [s for s in schools if _school_key(s) not in completed]

    if args.limit is not None:
        pending = pending[: args.limit]

    print("=" * 72)
    print("Schema-driven crawler batch (from spreadsheet)")
    print(f"  mode                  : {'retry-under-explored' if args.retry_under_explored else 'full'}")
    print(f"  input                 : {input_path}")
    print(f"  results sheet         : {RESULTS_SHEET}")
    print(f"  json backup           : {args.json_out}")
    if args.retry_under_explored:
        print(f"  retry progress        : {args.retry_json_out}")
        print(f"  candidate threshold   : < {args.candidate_threshold}")
    print(f"  unique schools        : {len(schools)}")
    if args.retry_under_explored:
        print(f"  retry targets         : {len(retry_targets)}")
        print(f"  already retried       : {len(retried) if args.resume else 0}")
    else:
        print(f"  already done (resume) : {len(existing_results) if args.resume else 0}")
    print(f"  to crawl this run     : {len(pending)}")
    print(f"  max_pages             : {max_pages}")
    print(f"  max_depth             : {args.max_depth}")
    print(f"  skip_archival         : {args.skip_archival}")
    print(f"  concurrency           : {args.concurrency}")
    print("=" * 72)

    if not pending:
        print("Nothing to crawl.")
        if args.json_out.exists():
            data = json.loads(args.json_out.read_text(encoding="utf-8"))
            write_results_sheet(input_path, data.get("districts", []))
            print(f"Refreshed {RESULTS_SHEET} from existing JSON.")
        return

    if not args.retry_under_explored and not existing_results:
        existing_results = []

    results: list[dict[str, Any]] = list(existing_results)
    retried_results: list[dict[str, Any]] = (
        load_results(args.retry_json_out) if args.retry_under_explored and args.resume else []
    )

    sem = asyncio.Semaphore(args.concurrency)
    classifier = PageClassifier()
    checkpoint_lock = asyncio.Lock()
    done_count = len(retried_results) if args.retry_under_explored else len(existing_results)
    run_total = len(retry_targets) if args.retry_under_explored else len(schools)
    completed_this_run = 0

    async def _worker(record: dict[str, str]) -> dict[str, Any]:
        async with sem:
            res = await crawl_one(
                record,
                classifier,
                max_depth=args.max_depth,
                skip_archival=args.skip_archival,
                max_pages=max_pages,
                max_candidates=max_candidates,
            )
            if args.retry_under_explored:
                res["retry_pass"] = True
                res["retry_max_pages"] = max_pages
            status = res["status"]
            budget = "BUDGET" if res.get("max_pages_limit_reached") else "ok"
            print(
                f"  [{status}] {res['name']:<42} "
                f"-> {res.get('candidate_count', 0)} candidates, "
                f"{res.get('pages_crawled', 0)} pages, "
                f"budget_reached={budget}"
                + (f"  ({res.get('error')})" if res.get("error") else "")
            )
            return res

    async def _process_one(record: dict[str, str]) -> None:
        nonlocal done_count, results, completed_this_run
        res = await _worker(record)
        async with checkpoint_lock:
            completed_this_run += 1
            if args.retry_under_explored:
                results = merge_result_update(results, res, order=order)
                retried_results.append(res)
                write_json_report(
                    args.retry_json_out,
                    retried_results,
                    meta={
                        "mode": "retry-under-explored",
                        "source_json": str(args.json_out),
                        "max_pages": max_pages,
                        "candidate_threshold": args.candidate_threshold,
                        "retried": len(retried_results),
                        "retry_targets": len(retry_targets),
                    },
                )
            else:
                results.append(res)
                results = sorted(
                    results,
                    key=lambda r: order.get(
                        _school_key(
                            {"name": r.get("name", ""), "website": r.get("website", "")}
                        ),
                        10_000,
                    ),
                )
            write_json_report(
                args.json_out,
                results,
                meta={
                    "input": str(input_path),
                    "mode": "retry-under-explored" if args.retry_under_explored else "full",
                    "max_pages": max_pages,
                    "max_depth": args.max_depth,
                    "skip_archival": args.skip_archival,
                    "processed": len(results),
                    "total_unique": len(schools),
                },
            )
            write_results_sheet(input_path, results)
            done_count = len(retried_results) if args.retry_under_explored else len(results)
            print(
                f"  checkpoint {done_count}/{run_total} "
                f"({completed_this_run}/{len(pending)} this run) "
                f"-> sheet + JSON updated"
            )

    await asyncio.gather(*[_process_one(record) for record in pending])

    ok = sum(1 for r in results if r.get("status") == "ok")
    budget_hits = sum(1 for r in results if r.get("max_pages_limit_reached"))
    total_candidates = sum(r.get("candidate_count", 0) for r in results)

    print("\n" + "=" * 72)
    print("Summary")
    print("=" * 72)
    print(f"  schools in sheet      : {len(results)}")
    print(f"  ok                    : {ok}")
    print(f"  max_pages_limit hits  : {budget_hits}")
    print(f"  total candidates      : {total_candidates}")
    print(f"\nResults sheet : {input_path} [{RESULTS_SHEET}]")
    print(f"JSON backup   : {args.json_out}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Run schema-driven crawler over schools in a spreadsheet."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument(
        "--json-out",
        type=Path,
        default=DEFAULT_JSON_OUT,
        help="Incremental JSON backup (also used by --resume).",
    )
    parser.add_argument("--max-depth", type=int, default=4)
    parser.add_argument(
        "--max-pages",
        type=int,
        default=None,
        help="Override LLM page budget (default: 15 full run, 30 retry run).",
    )
    parser.add_argument(
        "--retry-under-explored",
        action="store_true",
        help="Re-crawl schools that hit the page budget with fewer than "
        f"{DEFAULT_RETRY_CANDIDATE_THRESHOLD} candidates.",
    )
    parser.add_argument(
        "--candidate-threshold",
        type=int,
        default=DEFAULT_RETRY_CANDIDATE_THRESHOLD,
        help="Retry schools with fewer than this many candidates when "
        "--retry-under-explored is set.",
    )
    parser.add_argument(
        "--retry-json-out",
        type=Path,
        default=DEFAULT_RETRY_JSON_OUT,
        help="Retry progress JSON used by --resume during --retry-under-explored.",
    )
    parser.add_argument(
        "--skip-archival",
        action=argparse.BooleanOptionalAction,
        default=False,
        help="When True, skip pages the LLM marks is_archive. Default False.",
    )
    parser.add_argument("--concurrency", type=int, default=6)
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument(
        "--resume",
        action="store_true",
        help="Skip schools already present in the JSON backup.",
    )
    args = parser.parse_args()

    if args.max_depth < 0 or args.max_depth > 4:
        print("--max-depth must be between 0 and 4.", file=sys.stderr)
        sys.exit(1)
    if args.concurrency < 1:
        print("--concurrency must be >= 1.", file=sys.stderr)
        sys.exit(1)

    try:
        asyncio.run(run(args))
    except KeyboardInterrupt:
        print("\nInterrupted — partial results saved to JSON + results sheet.", file=sys.stderr)
        sys.exit(130)
    except Exception as exc:  # noqa: BLE001
        print(f"\nBatch crawl failed: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
