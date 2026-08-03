#!/usr/bin/env python3
"""
Seed the `charter_district_mapping` table from a user-provided spreadsheet.

Maps each charter school (by org_code) to its parent public school district
(also by org_code). Idempotent on (tenant_id, charter_org_code): existing
rows are updated in place; new rows are inserted; rows missing from the
spreadsheet are optionally removed with --remove-missing.

The expected spreadsheet format is intentionally flexible. The script looks
for these column-name patterns (case-insensitive, substring match):

  Charter org-code column:  'charter' AND 'org_code' OR 'charter_org_code'
  Parent org-code column:   'parent' AND 'org_code' OR 'parent_district_org_code'
                            OR 'public_org_code' OR 'district_org_code'

If your spreadsheet uses different headers, override with --charter-col and
--parent-col. The org codes are stripped and zero-padded to 8 characters
to match the schools.org_code format from DESE source data.

Usage:
    python scripts/school_data/seed_charter_mapping.py
    python scripts/school_data/seed_charter_mapping.py --xlsx path/to/mapping.xlsx
    python scripts/school_data/seed_charter_mapping.py --csv path/to/mapping.csv
    python scripts/school_data/seed_charter_mapping.py --remove-missing
    python scripts/school_data/seed_charter_mapping.py --dry-run

Requires schools to be seeded first (seed_schools.py) so the composite FK
constraints on (tenant_id, org_code) pass.
"""

import argparse
import asyncio
import sys
from pathlib import Path
from typing import Any

from sqlalchemy import delete, select
from sqlalchemy.exc import IntegrityError

from app.core.config import settings
from app.db.connector import AsyncSessionLocal
from app.models.charter_district_mapping import CharterDistrictMapping
from app.models.school import School

DEFAULT_EXCEL_PATH = (
    Path(__file__).parent / "output" / "charter_district_mapping.xlsx"
)

# Column-name detection patterns. Substrings matched case-insensitively.
_CHARTER_PATTERNS = (
    ("charter_org_code", "exact"),
    ("charter_org", "substring"),
    ("charter", "substring"),  # fallback: any 'charter' col
)
_PARENT_PATTERNS = (
    ("parent_district_org_code", "exact"),
    ("parent_org_code", "exact"),
    ("public_org_code", "exact"),
    ("district_org_code", "exact"),
    ("parent_org", "substring"),
    ("parent", "substring"),
)


def _norm_org_code(raw: Any) -> str | None:
    """Normalize an org code to a stripped 8-char zero-padded string."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    # DESE org codes are 8-digit numeric strings. Preserve non-numeric
    # codes (for non-MA states) but zero-pad pure numeric ones to 8 chars.
    if s.isdigit() and len(s) <= 8:
        s = s.zfill(8)
    return s


def _match_header(headers: list[str], patterns: tuple[tuple[str, str], ...]) -> str | None:
    """Find the first header matching any pattern (case-insensitive)."""
    lower = [h.strip().lower() for h in headers]
    for needle, kind in patterns:
        if kind == "exact":
            for h in lower:
                if h == needle:
                    return h
        else:  # substring
            for h in lower:
                if needle in h:
                    return h
    return None


def load_mapping_rows(
    path: Path,
    *,
    charter_col: str | None = None,
    parent_col: str | None = None,
) -> list[tuple[str, str]]:
    """
    Load the spreadsheet and return a list of (charter_org_code, parent_org_code)
    tuples. Auto-detects the charter and parent columns unless overridden.
    """
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "openpyxl is required to read .xlsx files. "
                "Run `poetry add openpyxl` or use --csv instead."
            ) from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    elif suffix == ".csv":
        import csv

        with path.open("r", encoding="utf-8-sig") as fh:
            rows = list(csv.reader(fh))
    else:
        raise ValueError(
            f"Unsupported file extension {suffix!r}. Use .xlsx or .csv."
        )

    if not rows:
        raise ValueError(f"Spreadsheet {path} is empty.")

    headers = [str(c) if c is not None else "" for c in rows[0]]
    c_col = charter_col or _match_header(headers, _CHARTER_PATTERNS)
    p_col = parent_col or _match_header(headers, _PARENT_PATTERNS)
    if not c_col:
        raise ValueError(
            "Could not auto-detect the charter org_code column. "
            f"Found headers: {headers}. Pass --charter-col to specify."
        )
    if not p_col:
        raise ValueError(
            "Could not auto-detect the parent org_code column. "
            f"Found headers: {headers}. Pass --parent-col to specify."
        )

    c_idx = headers.index(c_col)
    p_idx = headers.index(p_col)

    out: list[tuple[str, str]] = []
    for row in rows[1:]:
        if row is None:
            continue
        c = _norm_org_code(row[c_idx] if c_idx < len(row) else None)
        p = _norm_org_code(row[p_idx] if p_idx < len(row) else None)
        if c and p:
            out.append((c, p))
    return out


async def seed_mapping(
    rows: list[tuple[str, str]],
    *,
    tenant_id: int,
    remove_missing: bool = False,
    dry_run: bool = False,
) -> dict:
    """Upsert mapping rows into charter_district_mapping."""
    print("=" * 60)
    print("Just-EdTech Charter District Mapping Seeder")
    print(f"  tenant_id         : {tenant_id}")
    print(f"  rows to load      : {len(rows)}")
    print(f"  remove_missing    : {remove_missing}")
    print(f"  dry_run           : {dry_run}")
    print("=" * 60)

    stats = {
        "inserted": 0,
        "updated": 0,
        "unchanged": 0,
        "skipped_self": 0,
        "skipped_orphan_charter": 0,
        "skipped_orphan_parent": 0,
        "removed": 0,
    }
    seen_charter: set[str] = set()

    async with AsyncSessionLocal() as db:
        # Pre-load existing org_codes from schools so we can skip orphans
        # without N+1 queries.
        all_schools = (
            await db.execute(
                select(School.org_code).where(School.tenant_id == tenant_id)
            )
        ).scalars().all()
        school_orgs: set[str] = {oc for oc in all_schools if oc}
        if not school_orgs:
            print(
                "\nWARNING: schools table is empty. Run "
                "`python scripts/school_data/seed_schools.py` first.\n"
            )

        for charter_org, parent_org in rows:
            if charter_org == parent_org:
                stats["skipped_self"] += 1
                continue
            if charter_org not in school_orgs:
                stats["skipped_orphan_charter"] += 1
                continue
            if parent_org not in school_orgs:
                stats["skipped_orphan_parent"] += 1
                continue
            seen_charter.add(charter_org)

            existing = (
                await db.execute(
                    select(CharterDistrictMapping).where(
                        CharterDistrictMapping.tenant_id == tenant_id,
                        CharterDistrictMapping.charter_org_code == charter_org,
                    )
                )
            ).scalar_one_or_none()

            if existing is None:
                db.add(
                    CharterDistrictMapping(
                        tenant_id=tenant_id,
                        charter_org_code=charter_org,
                        parent_district_org_code=parent_org,
                    )
                )
                stats["inserted"] += 1
            elif existing.parent_district_org_code != parent_org:
                existing.parent_district_org_code = parent_org
                stats["updated"] += 1
            else:
                stats["unchanged"] += 1

        if remove_missing and not dry_run:
            result = await db.execute(
                delete(CharterDistrictMapping).where(
                    CharterDistrictMapping.tenant_id == tenant_id,
                    ~CharterDistrictMapping.charter_org_code.in_(seen_charter),
                )
            )
            stats["removed"] = int(result.rowcount or 0)

        if dry_run:
            await db.rollback()
            print("\n[dry-run] Changes NOT committed.")
        else:
            try:
                await db.commit()
            except IntegrityError as exc:
                await db.rollback()
                print(
                    f"\nIntegrity error (likely a FK violation because the "
                    f"org_code is not in the schools table): {exc.orig}",
                    file=sys.stderr,
                )
                sys.exit(2)

    print("\nSeed results:")
    for k, v in stats.items():
        print(f"  {k:<24}: {v}")
    print("\nDone.")
    return stats


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Seed charter_district_mapping from a spreadsheet."
    )
    src = parser.add_mutually_exclusive_group()
    src.add_argument(
        "--xlsx",
        type=Path,
        default=DEFAULT_EXCEL_PATH,
        help="Path to the mapping spreadsheet (.xlsx or .csv).",
    )
    src.add_argument(
        "--csv",
        type=Path,
        default=None,
        help="Path to a CSV mapping file (alternative to --xlsx).",
    )
    parser.add_argument(
        "--charter-col",
        type=str,
        default=None,
        help="Override the charter org_code column header.",
    )
    parser.add_argument(
        "--parent-col",
        type=str,
        default=None,
        help="Override the parent org_code column header.",
    )
    parser.add_argument(
        "--tenant-id",
        type=int,
        default=int(getattr(settings, "DEFAULT_TENANT_ID", 1) or 1),
        help="Tenant ID to scope the mapping (default: settings.DEFAULT_TENANT_ID).",
    )
    parser.add_argument(
        "--remove-missing",
        action="store_true",
        help="Remove mapping rows absent from the spreadsheet.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Validate and print stats without committing.",
    )
    args = parser.parse_args()

    path = args.csv or args.xlsx
    if not path.exists():
        print(
            f"\nMapping file not found: {path}\n"
            f"Drop the Excel/CSV into scripts/school_data/ and pass --xlsx "
            f"or --csv explicitly.\n",
            file=sys.stderr,
        )
        sys.exit(1)

    try:
        rows = load_mapping_rows(
            path,
            charter_col=args.charter_col,
            parent_col=args.parent_col,
        )
    except (ValueError, RuntimeError) as exc:
        print(f"\nFailed to parse mapping file: {exc}", file=sys.stderr)
        sys.exit(1)

    try:
        asyncio.run(
            seed_mapping(
                rows,
                tenant_id=args.tenant_id,
                remove_missing=args.remove_missing,
                dry_run=args.dry_run,
            )
        )
    except Exception as exc:
        print(f"\nSeeding failed: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
