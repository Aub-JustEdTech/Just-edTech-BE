"""XLSX / XLS document processor with table-aware chunking."""

import logging
from typing import Any

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

from app.services.document_processing.base import DocumentProcessor
from app.services.document_processing.markdown_table import (
    cell_value as _raw_cell_value,
)
from app.services.document_processing.markdown_table import (
    make_separator as _make_separator,
)
from app.services.document_processing.markdown_table import (
    row_to_markdown as _row_to_markdown,
)

logger = logging.getLogger(__name__)

# How many data rows to include in each chunk (headers are always repeated).
_DEFAULT_ROWS_PER_CHUNK = 25

# Char cap alongside the row cap: a wide sheet (many columns) can otherwise
# produce an oversized chunk despite staying under the row limit.
_DEFAULT_MAX_CHUNK_CHARS = 4000

# openpyxl loads .xls as well via xlrd compatibility, but we guard the import.
_SUPPORTED_EXTENSIONS = [".xlsx", ".xls"]
_SUPPORTED_MIME_TYPES = [
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
]


def _cell_value(cell) -> str:
    """Coerce an openpyxl cell to a clean string safe for Markdown tables."""
    return _raw_cell_value(cell.value)


def _find_header_row_index(sheet) -> int:
    """Return the 0-based index of the first row that has at least two non-empty cells."""
    for idx, row in enumerate(sheet.iter_rows()):
        non_empty = sum(1 for c in row if c.value not in (None, ""))
        if non_empty >= 2:
            return idx
    return 0


def _sheet_to_markdown(sheet) -> tuple[str, list[str], list[list[str]]]:
    """
    Convert one worksheet into:
        - full_markdown  : the complete markdown table string
        - headers        : list of header strings
        - data_rows      : list of rows, each a list of cell strings

    Title-like rows above the header (single non-empty cell across the full
    width or merged-cell spans) are prepended as plain text above the table.
    """
    all_rows = list(sheet.iter_rows())
    if not all_rows:
        return "", [], []

    header_idx = _find_header_row_index(sheet)

    # Collect any title text sitting above the header row.
    title_lines: list[str] = []
    for row in all_rows[:header_idx]:
        values = [_cell_value(c) for c in row]
        non_empty = [v for v in values if v]
        if non_empty:
            title_lines.append(" ".join(non_empty))

    header_row = all_rows[header_idx]
    headers = [_cell_value(c) for c in header_row]
    col_count = len(headers)

    data_rows: list[list[str]] = []
    for row in all_rows[header_idx + 1 :]:
        values = [_cell_value(c) for c in row]
        # Pad / trim to match header column count.
        values = (values + [""] * col_count)[:col_count]
        # Skip completely empty rows.
        if any(v for v in values):
            data_rows.append(values)

    lines: list[str] = []
    if title_lines:
        lines.extend(title_lines)
        lines.append("")

    lines.append(_row_to_markdown(headers))
    lines.append(_make_separator(col_count))
    for row in data_rows:
        lines.append(_row_to_markdown(row))

    return "\n".join(lines), headers, data_rows


class XLSXProcessor(DocumentProcessor):
    """Process Excel workbooks (.xlsx / .xls) into Markdown table text."""

    supported_extensions = _SUPPORTED_EXTENSIONS
    supported_mime_types = _SUPPORTED_MIME_TYPES

    # ------------------------------------------------------------------
    # Public interface (DocumentProcessor ABC)
    # ------------------------------------------------------------------

    def extract_text(self, file_path: str) -> str:
        """
        Return the full workbook as Markdown tables, one section per sheet.

        This is used by the summarizer pipeline stage.  The text preserves
        every sheet's header row and all data rows so the LLM can read the
        complete document.
        """
        self._require_openpyxl()
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        sections: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            table_md, _, _ = _sheet_to_markdown(ws)
            if table_md.strip():
                sections.append(f"## Sheet: {sheet_name}\n\n{table_md}")

        wb.close()

        if not sections:
            return ""

        full_text = "\n\n---\n\n".join(sections)
        logger.info(
            f"Extracted {len(full_text)} chars from {len(sections)} sheet(s) in {file_path}"
        )
        return full_text

    def extract_metadata(self, file_path: str) -> dict[str, Any]:
        """Return workbook-level metadata (sheet names, row/col counts)."""
        self._require_openpyxl()
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheets_info: list[dict[str, Any]] = []
            total_rows = 0
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                row_count = ws.max_row or 0
                col_count = ws.max_column or 0
                total_rows += row_count
                sheets_info.append(
                    {
                        "sheet_name": sheet_name,
                        "row_count": row_count,
                        "col_count": col_count,
                    }
                )
            wb.close()
            return {
                "sheet_count": len(sheets_info),
                "sheets": sheets_info,
                "total_row_count": total_rows,
                "document_format": "xlsx",
            }
        except Exception as exc:
            logger.error(f"Error extracting XLSX metadata from {file_path}: {exc}")
            return {"sheet_count": 0, "sheets": [], "total_row_count": 0}

    def validate(self, file_path: str) -> bool:
        """Return True if the file can be opened as a workbook."""
        if openpyxl is None:
            return False
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            wb.close()
            return True
        except Exception:
            return False

    # ------------------------------------------------------------------
    # XLSX-specific chunking
    # ------------------------------------------------------------------

    def chunk_spreadsheet(
        self,
        file_path: str,
        rows_per_chunk: int = _DEFAULT_ROWS_PER_CHUNK,
        max_chunk_chars: int = _DEFAULT_MAX_CHUNK_CHARS,
    ) -> list[dict[str, Any]]:
        """
        Chunk the workbook into row-group segments, repeating column headers
        at the top of every chunk.

        Returns a list of dicts:
            {
                "text":       <markdown string>,
                "sheet_name": <worksheet name>,
                "row_start":  <1-based first data row index in this chunk>,
                "row_end":    <1-based last data row index in this chunk>,
            }

        Strategy
        --------
        - Each sheet is treated independently.
        - The header row is detected automatically and repeated in every chunk.
        - Data rows are grouped into batches, closing a batch when EITHER
          `rows_per_chunk` rows or `max_chunk_chars` characters is reached —
          a wide sheet (many columns) would otherwise produce an oversized
          chunk despite staying under the row cap.
        - Small sheets (total data rows <= rows_per_chunk, under the char
          cap) produce a single chunk.
        - Entirely empty sheets are skipped.
        """
        self._require_openpyxl()
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        chunks: list[dict[str, Any]] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            _, headers, data_rows = _sheet_to_markdown(ws)

            if not headers or not data_rows:
                logger.debug(f"Skipping empty sheet '{sheet_name}' in {file_path}")
                continue

            col_count = len(headers)
            header_line = _row_to_markdown(headers)
            separator_line = _make_separator(col_count)
            base_chars = len(header_line) + len(separator_line) + 2  # + newlines

            batch: list[list[str]] = []
            batch_chars = base_chars
            batch_start = 0  # 0-based index into data_rows

            def _flush(
                batch,
                batch_start_idx,
                *,
                sheet_name=sheet_name,
                header_line=header_line,
                separator_line=separator_line,
                total_rows=len(data_rows),
            ):
                row_start = batch_start_idx + 1  # 1-based
                row_end = batch_start_idx + len(batch)  # 1-based, inclusive
                if row_start == 1 and row_end == total_rows:
                    heading = f"## Sheet: {sheet_name}"
                else:
                    heading = f"## Sheet: {sheet_name} (rows {row_start}–{row_end})"
                lines = [heading, "", header_line, separator_line]
                lines.extend(_row_to_markdown(row) for row in batch)
                return {
                    "text": "\n".join(lines),
                    "sheet_name": sheet_name,
                    "row_start": row_start,
                    "row_end": row_end,
                }

            for i, row in enumerate(data_rows):
                row_line = _row_to_markdown(row)
                row_chars = len(row_line) + 1  # + newline
                would_exceed_rows = len(batch) >= rows_per_chunk
                would_exceed_chars = batch and (
                    batch_chars + row_chars > max_chunk_chars
                )
                if batch and (would_exceed_rows or would_exceed_chars):
                    chunks.append(_flush(batch, batch_start))
                    batch = []
                    batch_chars = base_chars
                    batch_start = i

                batch.append(row)
                batch_chars += row_chars

            if batch:
                chunks.append(_flush(batch, batch_start))

        wb.close()
        logger.info(f"Produced {len(chunks)} spreadsheet chunk(s) from {file_path}")
        return chunks

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _require_openpyxl() -> None:
        if openpyxl is None:
            raise ImportError(
                "openpyxl is required for XLSX processing. "
                "Install it with: pip install openpyxl"
            )
